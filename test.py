import random
from openpyxl import load_workbook

# 打开原始Excel文件
workbook = load_workbook('班级民主测评统计表.xlsx')

# 选择要修改的工作表
sheet = workbook['Sheet1']

# 指定要填写的范围
start_row = 20
end_row = 20
start_col = 2  # 列B
end_col = 150  # 列DY

# 填写范围内的每个单元格为随机的8或10
for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        random_value = random.choice([8, 10])
        sheet.cell(row=row, column=col, value=random_value)

# 保存修改后的文件
workbook.save('班级民主测评统计表.xlsx')
